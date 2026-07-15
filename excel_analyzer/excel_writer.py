import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# Define color hexes
HEX_GREEN = "C6EFCE"        # Light green
HEX_RED = "FFC7CE"          # Light red
HEX_YELLOW = "FFEB9C"       # Light yellow

# Fonts: all normal, size 10, black text
font_normal = Font(name="Segoe UI", size=10, bold=False, color="000000")
font_bold = Font(name="Segoe UI", size=10, bold=False, color="000000")
font_header = Font(name="Segoe UI", size=10, bold=False, color="000000")
font_bold_10 = Font(name="Segoe UI", size=10, bold=False, color="000000")

# Color Fills: Light Blue 2 is DDEBF7, Light Yellow 2 is FFF2CC
fill_header = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
fill_summary = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
fill_overall = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
fill_sub_header = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
fill_header7 = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

fill_green = PatternFill(start_color=HEX_GREEN, end_color=HEX_GREEN, fill_type="solid")
fill_red = PatternFill(start_color=HEX_RED, end_color=HEX_RED, fill_type="solid")
fill_yellow = PatternFill(start_color=HEX_YELLOW, end_color=HEX_YELLOW, fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
)

def auto_fit_columns(ws, min_width=12):
    """Adjust column widths dynamically."""
    for col in ws.columns:
        max_len = 0
        for cell in col:
            # handle percentages nicely for width estimation
            if cell.number_format == "0.00%" and isinstance(cell.value, float):
                val_str = f"{cell.value * 100:.2f}%"
            else:
                val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, min_width)

def write_compact_summary_sheet(ws, summary_data, overall_data):
    ws.views.sheetView[0].showGridLines = True
    
    headers = [
        "Field Name", "Cumulative Correct", "Correct", "Correct (ND)", "Inferred",
        "Missing", "Incorrect", "Hallucination", "Total calls", "Average"
    ]
    ws.append(headers)
    
    for r in summary_data:
        ws.append([
            r["Field Name"],
            r["Cumulative Correct"],
            r["Correct"],
            r["Correct (ND)"],
            r["Inferred"],
            r["Missing"],
            r["Incorrect"],
            r["Hallucination"],
            r["Total Calls"],
            r["Average"]
        ])
        
    # Total calls fields = yy (using '=' instead of '==')
    overall_label = f"Total correct = {overall_data['total_correct']} Total calls fields = {overall_data['total_calls_fields']} , overall accuracy = {overall_data['total_correct']}/{overall_data['total_calls_fields']}"
    
    ws.append([
        overall_label,
        "", "", "", "", "", "", "", "",
        overall_data["overall_accuracy"]
    ])
    
    # Styles
    for cell in ws[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    num_rows = len(summary_data)
    for row_idx in range(2, num_rows + 2):
        row_cells = ws[row_idx]
        row_cells[0].font = font_normal
        row_cells[0].alignment = Alignment(horizontal="left", vertical="center")
        row_cells[0].border = thin_border
        
        for col_idx in range(1, 9):
            cell = row_cells[col_idx]
            cell.font = font_normal
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = thin_border
            cell.number_format = "#,##0"
            
        acc_cell = row_cells[9]
        acc_cell.font = font_normal
        acc_cell.alignment = Alignment(horizontal="right", vertical="center")
        acc_cell.border = thin_border
        acc_cell.number_format = "0.00%"
        
    # Overall summary row
    summary_r_idx = num_rows + 2
    for cell in ws[summary_r_idx]:
        cell.font = font_bold
        cell.fill = fill_overall
        cell.border = thin_border
    
    ws.cell(row=summary_r_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")
    acc_cell = ws.cell(row=summary_r_idx, column=10)
    acc_cell.alignment = Alignment(horizontal="right", vertical="center")
    acc_cell.number_format = "0.00%"
    
    ws.merge_cells(start_row=summary_r_idx, start_column=1, end_row=summary_r_idx, end_column=9)
    
    ws.freeze_panes = "A2"
    auto_fit_columns(ws)

def write_compact_overview(ws, overview_data, summary_ai, summary_manual):
    ws.views.sheetView[0].showGridLines = True
    
    # helper for styling tables
    def style_block_header(ws, start_row, end_col, title):
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=end_col)
        cell = ws.cell(row=start_row, column=1, value=title)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
    # Table 1: PNS Call Summary AI Audit
    style_block_header(ws, 1, 2, "1. PNS Call Summary AI Audit")
    ws.append(["Metric", "Value"])
    
    yes_ai = overview_data["ai_summary"]["yes"]
    total_ai = overview_data["ai_summary"]["total"]
    no_ai = overview_data["ai_summary"]["no"]
    acc_ai_val = yes_ai / total_ai if total_ai > 0 else 0.0
    acc_ai_str = f"{yes_ai}/{total_ai} | {acc_ai_val*100:.2f}%"
    
    ws.append(["Total calls AI audited", total_ai])
    ws.append(["No. of YES", yes_ai])
    ws.append(["No. of NO", no_ai])
    ws.append(["Accuracy", acc_ai_str])
    
    # Table 2: Manual Audit Summary
    ws.append([]) # spacer
    style_block_header(ws, 8, 2, "2. Manual Audit Summary")
    ws.append(["Metric", "Value"])
    
    yes_man = overview_data["manual_summary"]["yes"]
    total_man = overview_data["manual_summary"]["total"]
    no_man = overview_data["manual_summary"]["no"]
    acc_man_val = yes_man / total_man if total_man > 0 else 0.0
    acc_man_str = f"{yes_man}/{total_man} | {acc_man_val*100:.2f}%"
    
    ws.append(["Total calls manual audited", total_man])
    ws.append(["No. of YES", yes_man])
    ws.append(["No. of NO", no_man])
    ws.append(["Accuracy", acc_man_str])
    
    # Table 3: Field Wise Accuracy Comparison (5 columns: Denominator and % split)
    ws.append([]) # spacer
    style_block_header(ws, 15, 5, "3. Field Wise Accuracy Comparison")
    ws.append(["Field Name", "AI Audit Accuracy", "AI Audit Accuracy", "Manual Audit Accuracy", "Manual Audit Accuracy"])
    
    for r_ai, r_man in zip(summary_ai, summary_manual):
        f_name = r_ai["Field Name"]
        
        ai_cc = r_ai["Cumulative Correct"]
        ai_tot = r_ai["Total Calls"]
        ai_pct = ai_cc / ai_tot if ai_tot > 0 else 0.0
        
        man_cc = r_man["Cumulative Correct"]
        man_tot = r_man["Total Calls"]
        man_pct = man_cc / man_tot if man_tot > 0 else 0.0
        
        ws.append([
            f_name, 
            f"{ai_cc}/{ai_tot}", 
            ai_pct, 
            f"{man_cc}/{man_tot}", 
            man_pct
        ])
        
    # Table 4: AI v/s Manual Comparison Summary
    ws.append([]) # spacer
    style_block_header(ws, 31, 2, "4. AI v/s Manual Comparison Summary")
    ws.append(["Metric", "Value"])
    
    comp = overview_data["comparison_summary"]
    comp_acc = comp["matched_fields"] / comp["total_fields"] if comp["total_fields"] > 0 else 0.0
    comp_acc_str = f"{comp['matched_fields']}/{comp['total_fields']} | {comp_acc*100:.2f}%"
    
    # Matching exact requested labels
    ws.append(["Total call's matched field", comp["matched_fields"]])
    ws.append(["Total call's fields", comp["total_fields"]])
    ws.append(["Accuracy", comp_acc_str])
    
    # Styling Table 4 blocks
    for row_idx in range(1, ws.max_row + 1):
        if row_idx in [1, 8, 15, 31]:
            for cell in ws[row_idx]:
                cell.border = thin_border
            continue
            
        row_cells = ws[row_idx]
        if not any(c.value is not None for c in row_cells):
            continue # spacer row
            
        if row_idx in [2, 9, 16, 32]:
            for cell in row_cells:
                cell.font = font_bold
                cell.fill = fill_sub_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            continue
            
        # Format the 12 data rows in Table 3 (Rows 17 to 28)
        if 17 <= row_idx <= 28:
            row_cells[0].font = font_normal
            row_cells[0].alignment = Alignment(horizontal="left", vertical="center")
            row_cells[0].border = thin_border
            
            for col_i in [1, 3]: # Col 2 (AI den) and Col 4 (Man den)
                cell = row_cells[col_i]
                cell.font = font_normal
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                
            for col_i in [2, 4]: # Col 3 (AI %) and Col 5 (Man %)
                cell = row_cells[col_i]
                cell.font = font_normal
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = thin_border
                cell.number_format = "0.00%"
            continue
            
        # Table 1, 2, 4 general rows styling
        row_cells[0].font = font_normal
        row_cells[0].alignment = Alignment(horizontal="left", vertical="center")
        row_cells[0].border = thin_border
        
        for cell in row_cells[1:]:
            if cell.value is not None:
                cell.font = font_normal
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                
    auto_fit_columns(ws)

def create_excel_report(
    summary_ai, overall_ai,
    summary_manual, overall_manual,
    summary_ai_matched, overall_ai_matched,
    overview_data,
    data_calls, field_accuracies, overall_accuracy,
    output_path
):
    wb = openpyxl.Workbook()
    
    # 1. AI Audits Summary
    ws1 = wb.active
    ws1.title = "AI Audits Summary"
    write_compact_summary_sheet(ws1, summary_ai, overall_ai)
    
    # 2. Manual Audits Summary
    ws2 = wb.create_sheet(title="Manual Audits Summary")
    write_compact_summary_sheet(ws2, summary_manual, overall_manual)
    
    # 3. AI Matched Summary
    ws3 = wb.create_sheet(title="AI Matched Summary")
    write_compact_summary_sheet(ws3, summary_ai_matched, overall_ai_matched)
    
    # 4. Compact Overview
    ws4 = wb.create_sheet(title="Compact Overview")
    write_compact_overview(ws4, overview_data, summary_ai, summary_manual)
    
    # 5. Binary Comparison
    ws5 = wb.create_sheet(title="Binary Comparison")
    ws5.views.sheetView[0].showGridLines = True
    
    headers5 = [
        "Call ID", "Buyer Name", "Buyer Location", "Buyer Contact Details",
        "Lead Tag", "Product Name", "Specifications", "Buyer Required Quantity",
        "Price Quoted by Seller", "Seller Actionables", "Buyer Intent",
        "Buyer Questions", "Reminder", "Matches", "Total Fields", "Accuracy %"
    ]
    ws5.append(headers5)
    
    for item in data_calls:
        ws5.append([
            item["call_id"],
            item["matches_by_field"]["buyer_name"],
            item["matches_by_field"]["buyer_location"],
            item["matches_by_field"]["buyer_contact"],
            item["matches_by_field"]["lead_tag"],
            item["matches_by_field"]["product_name"],
            item["matches_by_field"]["specifications"],
            item["matches_by_field"]["quantity"],
            item["matches_by_field"]["price"],
            item["matches_by_field"]["actionables"],
            item["matches_by_field"]["buyer_intent"],
            item["matches_by_field"]["buyer_questions"],
            item["matches_by_field"]["reminder"],
            item["matches"],
            item["total_fields"],
            round(item["accuracy"], 2)
        ])
        
    summary_row5 = ["Overall Statistics"]
    for field_name in [
        "buyer_name", "buyer_location", "buyer_contact", "lead_tag",
        "product_name", "specifications", "quantity", "price",
        "actionables", "buyer_intent", "buyer_questions", "reminder"
    ]:
        acc = field_accuracies.get(field_name, 0.0)
        summary_row5.append(f"{acc:.2f}%")
    summary_row5.extend(["", "", f"Overall Accuracy: {overall_accuracy:.2f}%"])
    ws5.append(summary_row5)
    
    # Styles for Sheet 5
    ws5.freeze_panes = "A2"
    for cell in ws5[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    for r_idx in range(2, ws5.max_row):
        for cell in ws5[r_idx]:
            cell.font = font_normal
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
    summary_row_idx5 = ws5.max_row
    for cell in ws5[summary_row_idx5]:
        cell.font = font_bold
        cell.fill = fill_summary
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # Conditional formatting for B to M
    match_rule_1 = CellIsRule(operator='equal', formula=['1'], fill=fill_green)
    match_rule_0 = CellIsRule(operator='equal', formula=['0'], fill=fill_red)
    for col_idx in range(2, 14):
        col_letter = get_column_letter(col_idx)
        ws5.conditional_formatting.add(f"{col_letter}2:{col_letter}{summary_row_idx5-1}", match_rule_1)
        ws5.conditional_formatting.add(f"{col_letter}2:{col_letter}{summary_row_idx5-1}", match_rule_0)
        
    # Accuracy conditional formats
    acc_rule_green = CellIsRule(operator='greaterThanOrEqual', formula=['90'], fill=fill_green)
    acc_rule_yellow = CellIsRule(operator='between', formula=['75', '89.9'], fill=fill_yellow)
    acc_rule_red = CellIsRule(operator='lessThan', formula=['75'], fill=fill_red)
    ws5.conditional_formatting.add(f"P2:P{summary_row_idx5-1}", acc_rule_green)
    ws5.conditional_formatting.add(f"P2:P{summary_row_idx5-1}", acc_rule_yellow)
    ws5.conditional_formatting.add(f"P2:P{summary_row_idx5-1}", acc_rule_red)
    
    auto_fit_columns(ws5)
    
    # 6. Detailed Value Comparison
    ws6 = wb.create_sheet(title="Detailed Value Comparison")
    ws6.views.sheetView[0].showGridLines = True
    
    field_details = [
        ("buyer_name", "Buyer Name"),
        ("buyer_location", "Buyer Location"),
        ("buyer_contact", "Buyer Contact Details"),
        ("lead_tag", "Lead Tag"),
        ("product_name", "Product Name"),
        ("specifications", "Specifications"),
        ("quantity", "Buyer Required Quantity"),
        ("price", "Price Quoted by Seller"),
        ("actionables", "Seller Actionables"),
        ("buyer_intent", "Buyer Intent"),
        ("buyer_questions", "Buyer Questions"),
        ("reminder", "Reminder")
    ]
    
    headers6 = ["Call ID"]
    for _, field_label in field_details:
        headers6.append(f"{field_label} (M)")
        headers6.append(f"{field_label} (AI)")
    ws6.append(headers6)
    
    for item in data_calls:
        row_data = [item["call_id"]]
        for field_key, _ in field_details:
            m_val = item["manual_values"].get(field_key)
            a_val = item["ai_values"].get(field_key)
            row_data.append(m_val if m_val is not None else "")
            row_data.append(a_val if a_val is not None else "")
        ws6.append(row_data)
        
    ws6.freeze_panes = "A2"
    for cell in ws6[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    for r_idx in range(2, ws6.max_row + 1):
        ws6.cell(row=r_idx, column=1).font = font_bold
        for cell in ws6[r_idx]:
            if cell.column > 1:
                cell.font = font_normal
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
    auto_fit_columns(ws6)
    
    # 7. Summary Statistics (with Cumulative Correct column)
    ws7 = wb.create_sheet(title="Summary Statistics")
    ws7.views.sheetView[0].showGridLines = True
    
    headers7 = [
        "Field Name", "Cumulative Correct", "Correct", "Correct (Not Discussed)", "Inferred", "Incorrect",
        "Missing", "Hallucination", "Matched Calls", "Total calls", "Accuracy"
    ]
    ws7.append(headers7)
    
    total_calls_c = len(data_calls)
    field_rows_c = []
    
    for field_key, field_label in field_details:
        counts_c = {
            "Correct": 0,
            "Correct (Not Discussed)": 0,
            "Inferred": 0,
            "Incorrect": 0,
            "Missing": 0,
            "Hallucination": 0
        }
        
        matched_calls = 0
        for item in data_calls:
            val = item["manual_values"].get(field_key)
            if val in counts_c:
                counts_c[val] += 1
            matched_calls += item["matches_by_field"].get(field_key, 0)
            
        cum_correct = counts_c["Correct"] + counts_c["Correct (Not Discussed)"] + counts_c["Inferred"]
        acc_c = (matched_calls / total_calls_c) * 100 if total_calls_c > 0 else 0.0
        
        row_data = [
            field_label,
            cum_correct,
            counts_c["Correct"],
            counts_c["Correct (Not Discussed)"],
            counts_c["Inferred"],
            counts_c["Incorrect"],
            counts_c["Missing"],
            counts_c["Hallucination"],
            matched_calls,
            total_calls_c,
            acc_c
        ]
        ws7.append(row_data)
        field_rows_c.append(acc_c)
        
    overall_acc_c = sum(field_rows_c) / len(field_rows_c) if field_rows_c else 0.0
    ws7.append([
        "Overall Accuracy:", "", "", "", "", "", "", "", "", "", overall_acc_c
    ])
    
    # Styles for Sheet 7
    for cell in ws7[1]:
        cell.font = font_bold_10
        cell.fill = fill_header7
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    # Data rows styles
    for r_idx in range(2, ws7.max_row):
        ws7.cell(row=r_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws7.cell(row=r_idx, column=1).font = font_normal
        ws7.cell(row=r_idx, column=1).border = thin_border
        
        # 11 columns in total: index 2 to 10
        for c_idx in range(2, 11):
            cell = ws7.cell(row=r_idx, column=c_idx)
            cell.font = font_normal
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        ws7.cell(row=r_idx, column=11).number_format = '0"%"'
        
    summary_r_idx7 = ws7.max_row
    ws7.cell(row=summary_r_idx7, column=1).alignment = Alignment(horizontal="left", vertical="center")
    
    for cell in ws7[summary_r_idx7]:
        cell.font = font_bold_10
        cell.fill = fill_summary
        cell.border = thin_border
        if cell.column > 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
    ws7.cell(row=summary_r_idx7, column=11).number_format = '0.00"%"'
    ws7.freeze_panes = "A2"
    auto_fit_columns(ws7)
    
    # Global check: Apply thin border to every single cell with values in all worksheets
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = thin_border
                
    wb.save(output_path)
