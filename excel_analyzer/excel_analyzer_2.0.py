import os
import sys
import argparse
import logging
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure we can import from excel_analyzer.py in the same folder
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

try:
    from excel_analyzer import FIELD_TARGETS, FIELDS, normalize_value, map_columns
except ImportError:
    # Inline fallback mapping in case imports are tricky
    FIELDS = [
        "buyer_name", "buyer_location", "buyer_contact", "lead_tag",
        "product_name", "specifications", "quantity", "price",
        "actionables", "buyer_intent", "buyer_questions", "reminder"
    ]
    FIELD_TARGETS = {
        "buyer_name": {"buyername", "buyer name"},
        "buyer_location": {"buyerlocation", "buyer location"},
        "buyer_contact": {"buyercontact", "buyer contact", "buyer contact details", "contact"},
        "lead_tag": {"leadtag", "lead tag"},
        "product_name": {"productname", "product name"},
        "specifications": {"specifications", "specification"},
        "quantity": {"quantity", "buyerrequiredquantity", "buyer required quantity"},
        "price": {"price", "pricequotedbyseller", "price quoted by seller"},
        "actionables": {"actionables", "selleractionables", "seller actionables"},
        "buyer_intent": {"buyerintent", "buyer intent"},
        "buyer_questions": {"buyerquestions", "buyer questions"},
        "reminder": {"reminder"}
    }
    # Handled via local definitions if needed
    def normalize_badge(text: str) -> str:
        text = text.upper().strip()
        if any(k in text for k in ["INFERRED", "INFER", "INF"]): return "Inferred"
        if any(k in text for k in ["NOT DISCUSSED", "NDT DISCUSSED", "NOT DISCVSSED", "NDT DISCVSSED", "MISCVSSF"]): return "Correct (Not Discussed)"
        if any(k in text for k in ["INCORRECT", "INCO", "JNCO", "JNC", "INC", "NCR", "ICR"]): return "Incorrect"
        if any(k in text for k in ["CORRECT", "COAR", "COPP", "COER", "COPPECT", "COA", "COE", "CORR", "CDRRECT", "CDRR", "CDRE", "OPREIT", "OPRE", "IORRFIT", "IORR", "IRR", "ORR", "ORRFIT"]): return "Correct"
        if any(k in text for k in ["MISSING", "MISS", "MIS", "MSS", "MSI"]): return "Missing"
        if any(k in text for k in ["HALLUCINATION", "HALL", "HAL", "LUCI", "HAC"]): return "Hallucination"
        return None

    def normalize_value(val) -> str:
        if pd.isna(val) or val is None: return "Correct (Not Discussed)"
        val_str = str(val).strip()
        if val_str == "" or val_str.lower() in ["none", "null", "-", "nan"]: return "Correct (Not Discussed)"
        norm = normalize_badge(val_str)
        if norm: return norm
        val_upper = val_str.upper()
        if "NOT DISCUSSED" in val_upper: return "Correct (Not Discussed)"
        if "INFERRED" in val_upper: return "Inferred"
        if "INCORRECT" in val_upper: return "Incorrect"
        if "CORRECT" in val_upper: return "Correct"
        if "MISSING" in val_upper: return "Missing"
        if "HALLUCINATION" in val_upper: return "Hallucination"
        return "Correct"

    def map_columns(headers, fields_targets, sheet_name="Sheet"):
        mapping = {}
        normalized_headers = {}
        for h in headers:
            if not isinstance(h, str) and not isinstance(h, (int, float)): continue
            h_str = str(h)
            clean_h = re.sub(r'[^a-zA-Z0-9]', '', h_str).lower()
            if clean_h: normalized_headers[clean_h] = h_str
        call_id_col = None
        for target in {"callid", "call id"}:
            clean_t = re.sub(r'[^a-zA-Z0-9]', '', target).lower()
            if clean_t in normalized_headers:
                call_id_col = normalized_headers[clean_t]
                break
        if not call_id_col:
            for clean_h, orig_h in normalized_headers.items():
                if "callid" in clean_h:
                    call_id_col = orig_h
                    break
        mapping["call_id"] = call_id_col
        for field, targets in fields_targets.items():
            matched_header = None
            for t in targets:
                clean_t = re.sub(r'[^a-zA-Z0-9]', '', t).lower()
                if clean_t in normalized_headers:
                    matched_header = normalized_headers[clean_t]
                    break
            if matched_header:
                mapping[field] = matched_header
            else:
                for clean_h, orig_h in normalized_headers.items():
                    for t in targets:
                        clean_t = re.sub(r'[^a-zA-Z0-9]', '', t).lower()
                        if clean_t in clean_h and not any(sub in clean_h for sub in ["reason", "sublabel", "subfield"]):
                            matched_header = orig_h
                            break
                    if matched_header: break
                mapping[field] = matched_header
        return mapping

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

def create_summary_report(results, overall_row, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Field Summary"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    headers = [
        "Field Name", "Correct", "Correct (Not Discussed)", "Inferred",
        "Incorrect", "Missing", "Hallucination", "Total calls", "Accuracy"
    ]
    ws.append(headers)
    
    # Stylings
    font_normal = Font(name="Segoe UI", size=10)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_overall = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid") # soft amber/yellow for overall
    
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )
    
    # 1. Add rows
    for r in results:
        ws.append([
            r["Field Name"],
            r["Correct"],
            r["Correct (Not Discussed)"],
            r["Inferred"],
            r["Incorrect"],
            r["Missing"],
            r["Hallucination"],
            r["Total calls"],
            r["Accuracy"]
        ])
        
    # 2. Add overall row
    ws.append([
        overall_row["Field Name"],
        overall_row["Correct"],
        overall_row["Correct (Not Discussed)"],
        overall_row["Inferred"],
        overall_row["Incorrect"],
        overall_row["Missing"],
        overall_row["Hallucination"],
        overall_row["Total calls"],
        overall_row["Accuracy"]
    ])
    
    # 3. Apply styles
    # Header Row
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    # Data Rows
    num_rows = len(results)
    for row_idx in range(2, num_rows + 2):
        row_cells = ws[row_idx]
        # Field Name (left align)
        row_cells[0].font = font_normal
        row_cells[0].alignment = Alignment(horizontal="left", vertical="center")
        row_cells[0].border = thin_border
        
        # Counts (right align)
        for col_idx in range(1, 8):
            cell = row_cells[col_idx]
            cell.font = font_normal
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = thin_border
            cell.number_format = "#,##0"
            
        # Accuracy (right align, formatted as %)
        acc_cell = row_cells[8]
        acc_cell.font = font_normal
        acc_cell.alignment = Alignment(horizontal="right", vertical="center")
        acc_cell.border = thin_border
        acc_cell.number_format = "0.00%"
        
    # Overall Row (last row)
    overall_row_idx = num_rows + 2
    row_cells = ws[overall_row_idx]
    for cell in row_cells:
        cell.font = font_bold
        cell.fill = fill_overall
        cell.border = thin_border
        
    row_cells[0].alignment = Alignment(horizontal="left", vertical="center")
    
    for col_idx in range(1, 8):
        cell = row_cells[col_idx]
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if cell.value != "":
            cell.number_format = "#,##0"
            
    acc_cell = row_cells[8]
    acc_cell.alignment = Alignment(horizontal="right", vertical="center")
    acc_cell.number_format = "0.00%"
    
    # Autofit columns
    for col in ws.columns:
        max_len = 0
        for cell in col:
            # handle percentages nicely for width estimation
            if cell.number_format == "0.00%" and isinstance(cell.value, float):
                val_str = f"{cell.value * 100:.2f}%"
            else:
                val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(output_path)

def get_next_filename(base_dir, base_name="Compact_Analysis"):
    version = 1
    while True:
        candidate = os.path.join(base_dir, f"{base_name}_v{version}.xlsx")
        if not os.path.exists(candidate):
            return candidate
        version += 1

def main():
    parser = argparse.ArgumentParser(description="Excel Analyser 2.0 - Category count & accuracy reporting")
    parser.add_argument("--input-file", "-i", required=True, help="Path to input CSV file (AI Audits or Manual Audits)")
    parser.add_argument("--output-file", "-o", default=None, help="Path to save output Excel file")
    
    args = parser.parse_args()
    
    input_path = args.input_file
    if not os.path.exists(input_path):
        alternative_path = os.path.join(os.path.dirname(__file__), input_path)
        if os.path.exists(alternative_path):
            input_path = alternative_path
        else:
            logger.error(f"Input file '{input_path}' does not exist.")
            sys.exit(1)
            
    # Auto-generate output filename if not provided
    if not args.output_file:
        output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
        os.makedirs(output_dir, exist_ok=True)
        output_path = get_next_filename(output_dir, "Compact_Analysis")
    else:
        output_path = args.output_file
        
    logger.info(f"Loading data from: {input_path}")
    is_excel = input_path.lower().endswith(('.xlsx', '.xls'))
    try:
        if is_excel:
            df = pd.read_excel(input_path)
        else:
            try:
                df = pd.read_csv(input_path, encoding='utf-8')
            except UnicodeDecodeError:
                df = pd.read_csv(input_path, encoding='latin1')
    except Exception as e:
        file_type = "Excel" if is_excel else "CSV"
        logger.error(f"Failed to read {file_type} file '{input_path}': {str(e)}")
        sys.exit(1)
        
    # Map columns
    mapping = map_columns(df.columns, FIELD_TARGETS, "CSV")
    call_id_col = mapping.get("call_id")
    if not call_id_col:
        logger.error(f"Could not locate 'Call ID' column in CSV. Headers: {list(df.columns)}")
        sys.exit(1)
        
    # Clean and filter Call IDs
    df = df.dropna(subset=[call_id_col])
    
    # Locate Verdict column
    verdict_col = None
    for col in df.columns:
        if str(col).lower().strip() == 'verdict':
            verdict_col = col
            break
    if not verdict_col:
        for col in df.columns:
            if 'verdict' in str(col).lower():
                verdict_col = col
                break
                
    if not verdict_col:
        logger.error(f"Could not locate 'Verdict' column in CSV. Headers: {list(df.columns)}")
        sys.exit(1)
        
    # Filter dataset where Verdict is 'yes' or 'no' (case-insensitive)
    df_filtered = df[df[verdict_col].astype(str).str.strip().str.lower().isin(['yes', 'no'])]
    total_calls = len(df_filtered)
    
    logger.info(f"Filtered dataset from {len(df)} rows to {total_calls} rows where Verdict is 'yes' or 'no'.")
    if total_calls == 0:
        logger.error("No rows found where Verdict is 'yes' or 'no'. Exiting.")
        sys.exit(1)
        
    field_display_names = {
        "buyer_name": "Buyer Name",
        "buyer_location": "Buyer Location",
        "buyer_contact": "Buyer Contact Details",
        "lead_tag": "Lead Tag",
        "product_name": "Product Name",
        "specifications": "Specifications",
        "quantity": "Buyer Required Quantity",
        "price": "Price Quoted by Seller",
        "actionables": "Seller Actionables",
        "buyer_intent": "Buyer Intent",
        "buyer_questions": "Buyer Questions",
        "reminder": "Reminder"
    }

    results = []
    
    for field in FIELDS:
        display_name = field_display_names.get(field, field)
        col_name = mapping.get(field)
        
        counts = {
            "Correct": 0,
            "Correct (Not Discussed)": 0,
            "Inferred": 0,
            "Incorrect": 0,
            "Missing": 0,
            "Hallucination": 0
        }
        
        if not col_name or col_name not in df_filtered.columns:
            logger.warning(f"Field '{field}' was not mapped or is missing in CSV. Defaulting counts to 0.")
        else:
            for val in df_filtered[col_name]:
                norm_val = normalize_value(val)
                if norm_val in counts:
                    counts[norm_val] += 1
                else:
                    logger.warning(f"Unexpected normalized badge '{norm_val}' from '{val}'")
                    
        field_total = sum(counts.values())
        correct_sum = counts["Correct"] + counts["Correct (Not Discussed)"] + counts["Inferred"]
        accuracy = correct_sum / field_total if field_total > 0 else 0.0
        
        results.append({
            "Field Name": display_name,
            "Correct": counts["Correct"],
            "Correct (Not Discussed)": counts["Correct (Not Discussed)"],
            "Inferred": counts["Inferred"],
            "Incorrect": counts["Incorrect"],
            "Missing": counts["Missing"],
            "Hallucination": counts["Hallucination"],
            "Total calls": field_total,
            "Accuracy": accuracy
        })
        
    # Count how many calls have Verdict as 'yes' and 'no'
    yes_calls = sum(df_filtered[verdict_col].astype(str).str.strip().str.lower() == 'yes')
    no_calls = sum(df_filtered[verdict_col].astype(str).str.strip().str.lower() == 'no')
    overall_accuracy = yes_calls / total_calls if total_calls > 0 else 0.0
    
    overall_label = f"Overall accuracy ({yes_calls} was yes, {no_calls} was No and {total_calls} was total calls so {yes_calls}/{total_calls} is the overall accuracy)"
    
    overall_row = {
        "Field Name": overall_label,
        "Correct": "",
        "Correct (Not Discussed)": "",
        "Inferred": "",
        "Incorrect": "",
        "Missing": "",
        "Hallucination": "",
        "Total calls": total_calls,
        "Accuracy": overall_accuracy
    }
    
    logger.info(f"Generating summary report at: {output_path}")
    create_summary_report(results, overall_row, output_path)
    logger.info("Summary report generated successfully!")
    
    # Print a text table to the console for quick verification
    print("\n--- Summary Verification Table ---")
    print(f"{'Field Name':<25} | {'Corr':<5} | {'CorrND':<6} | {'Infer':<5} | {'Incorr':<6} | {'Miss':<4} | {'Hall':<4} | {'Total':<5} | {'Accuracy':<8}")
    print("-" * 88)
    for r in results:
        print(f"{r['Field Name']:<25} | {r['Correct']:<5} | {r['Correct (Not Discussed)']:<6} | {r['Inferred']:<5} | {r['Incorrect']:<6} | {r['Missing']:<4} | {r['Hallucination']:<4} | {r['Total calls']:<5} | {r['Accuracy']*100:.2f}%")
    print("-" * 88)
    print(f"{overall_row['Field Name']:<25} | {'':<5} | {'':<6} | {'':<5} | {'':<6} | {'':<4} | {'':<4} | {overall_row['Total calls']:<5} | {overall_row['Accuracy']*100:.2f}%")
    print("-" * 88)

if __name__ == "__main__":
    main()
