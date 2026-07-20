import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import config

def write_magicpen_report(input_rows: list, results: dict, output_path: str) -> None:
    """
    Creates the comparative and model metric sheets in an Excel workbook.
    
    Arguments:
        input_rows: list of dictionaries representing the original CSV input rows.
        results: dictionary mapping model_id -> list of query result dictionaries.
        output_path: file path to save the Excel workbook.
    """
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # Sheet 1: Comparison
    # ----------------------------------------------------
    ws_comp = wb.active
    ws_comp.title = "Comparison"
    
    # Build Comparison headers
    base_headers = [
        "Seller Glid", "Buyer Glid", "User Details", 
        "Product Details", "Past Conversation", "Ongoing Conversation"
    ]
    
    model_keys = list(config.MODELS_CONFIG.keys())
    model_headers = [config.MODELS_CONFIG[m]["excel_comparison_header"] for m in model_keys]
    comp_headers = base_headers + model_headers
    
    ws_comp.append(comp_headers)
    
    # Add data rows preserving original order
    for idx, row in enumerate(input_rows):
        row_data = [
            row.get("Seller Glid", ""),
            row.get("Buyer Glid", ""),
            row.get("User Details", ""),
            row.get("Product Details", ""),
            row.get("Past Conversation", ""),
            row.get("Ongoing Conversation", "")
        ]
        # Append response for each model corresponding to the current row
        for model in model_keys:
            model_res = results.get(model, [])
            if idx < len(model_res):
                row_data.append(model_res[idx]["response"])
            else:
                row_data.append("ERROR")
        ws_comp.append(row_data)
        
    # Apply styling to Comparison sheet
    style_sheet(ws_comp)
    
    # ----------------------------------------------------
    # Sheets 2-4: Model Metrics
    # ----------------------------------------------------
    for model in model_keys:
        model_conf = config.MODELS_CONFIG[model]
        sheet_name = model_conf["excel_sheet_name"]
        
        ws_metric = wb.create_sheet(title=sheet_name)
        
        # Build headers
        metric_headers = [
            "Call ID", "Prompt Tokens", "Completion Tokens", 
            "Total Tokens", "Estimated Cost", "Latency"
        ]
        ws_metric.append(metric_headers)
        
        model_res = results.get(model, [])
        for idx, res in enumerate(model_res):
            row_num = idx + 1
            ws_metric.append([
                row_num,
                res["prompt_tokens"],
                res["completion_tokens"],
                res["total_tokens"],
                res["cost"],
                res["latency"]
            ])
            
        # Format the data cells in the metrics sheet
        end_row = len(model_res) + 1
        for row in range(2, end_row + 1):
            # Prompt Tokens
            cell_prompt = ws_metric.cell(row=row, column=2)
            if isinstance(cell_prompt.value, (int, float)):
                cell_prompt.number_format = "#,##0"
            # Completion Tokens
            cell_comp = ws_metric.cell(row=row, column=3)
            if isinstance(cell_comp.value, (int, float)):
                cell_comp.number_format = "#,##0"
            # Total Tokens
            cell_total = ws_metric.cell(row=row, column=4)
            if isinstance(cell_total.value, (int, float)):
                cell_total.number_format = "#,##0"
            # Estimated Cost
            cell_cost = ws_metric.cell(row=row, column=5)
            if isinstance(cell_cost.value, (int, float)):
                cell_cost.number_format = "0.00000000"
            # Latency
            cell_lat = ws_metric.cell(row=row, column=6)
            if isinstance(cell_lat.value, (int, float)):
                cell_lat.number_format = "0.000"
                
        # Append summary metrics at the bottom using Excel formulas
        # Skip one blank row: blank is end_row + 1, summaries start at end_row + 2
        sum_start = end_row + 2
        
        summaries = [
            ("Average Prompt Tokens", f"=AVERAGE(B2:B{end_row})", 2, "#,##0"),
            ("Average Completion Tokens", f"=AVERAGE(C2:C{end_row})", 3, "#,##0"),
            ("Average Total Tokens", f"=AVERAGE(D2:D{end_row})", 4, "#,##0"),
            ("Average Estimated Cost", f"=AVERAGE(E2:E{end_row})", 5, "0.00000000"),
            ("Average Latency", f"=AVERAGE(F2:F{end_row})", 6, "0.000"),
            ("P95 Latency", f"=PERCENTILE.INC(F2:F{end_row},0.95)", 6, "0.000"),
            ("Maximum Latency", f"=MAX(F2:F{end_row})", 6, "0.000"),
            ("Minimum Latency", f"=MIN(F2:F{end_row})", 6, "0.000"),
            ("Standard Deviation Latency", f"=STDEV.S(F2:F{end_row})", 6, "0.000")
        ]
        
        bold_font = Font(name="Calibri", size=11, bold=True)
        
        for offset, (label, formula, col_idx, num_fmt) in enumerate(summaries):
            curr_row = sum_start + offset
            # Set label in column A
            cell_lbl = ws_metric.cell(row=curr_row, column=1, value=label)
            cell_lbl.font = bold_font
            
            # Set formula in target column
            cell_fml = ws_metric.cell(row=curr_row, column=col_idx, value=formula)
            cell_fml.number_format = num_fmt
            cell_fml.font = bold_font
            
        style_sheet(ws_metric)
        
    wb.save(output_path)

def style_sheet(ws) -> None:
    """
    Applies shared styling rules to a worksheet:
    - Bold headers
    - Freeze first row
    - Auto-adjust column widths
    """
    # 1. Bold headers & light gray fill
    header_font = Font(name="Calibri", size=11, bold=True, color="000000")
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
    # 2. Freeze first row
    ws.freeze_panes = "A2"
    
    # 3. Auto-adjust column widths
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            # Skip checking lengths of formulas or massive conversation values to prevent overly wide columns
            val = cell.value
            if val is not None and not str(val).startswith("="):
                # Restrict width calculation max check to 100 characters per cell check
                max_len = max(max_len, min(len(str(val)), 100))
        # Add padding
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
